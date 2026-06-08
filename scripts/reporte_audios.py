import sys
from pathlib import Path
from typing import Generator

# ── Configuración ──────────────────────────────────────────────────────────────
BASE_FOLDER  = Path(r"E:\ProcesoAudios\2026\example\01")  # ← carpeta a escanear
EXTENSIONES  = {".mp3", ".wav", ".ogg", ".flac", ".m4a"}  # tipos de audio
SALIDA_EXCEL = Path(r"E:\ProcesoAudios\2026\example\reporte_audios.xlsx")

# Excel permite 1 048 576 filas; reservamos 1 para cabecera → 1 048 575 datos
MAX_FILAS_POR_ARCHIVO = 1_048_575
# ──────────────────────────────────────────────────────────────────────────────

CAMPOS_FIJOS = ["YEAR", "COD1", "AoP", "COD2", "COD3", "TELEFONO"]

COLUMNAS_FIJAS: list[tuple[str, int]] = [
    ("MES",              8),
    ("DIA",             22),
    ("INDICE",           8),
    ("YEAR",            12),
    ("COD1",            10),
    ("AoP",              7),
    ("COD2",             8),
    ("COD3",             8),
    ("TELEFONO",        14),
    ("PESO",            12),
    ("RUTA",            60),
    ("NOMBRE_COMPLETO", 50),
]
NOMBRES_FIJOS = {c[0] for c in COLUMNAS_FIJAS}
COLUMNAS_CENTRADAS_FIJAS = {"MES", "INDICE", "YEAR", "COD1", "AoP", "COD2", "COD3", "PESO"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def formatear_peso(bytes_: int) -> str:
    return f"{bytes_ / 1024:.2f} KB"


def parsear_nombre(nombre_sin_ext: str) -> dict:
    """
    Divide el nombre por '-' y mapea posicionalmente a los 6 campos fijos.
    Bloques vacíos (guiones consecutivos) quedan como ''.
    Bloques 7+ se almacenan como BLOQUE_7, BLOQUE_8, ...
    """
    partes = nombre_sin_ext.split("-")
    resultado = {campo: (partes[i] if i < len(partes) else "")
                 for i, campo in enumerate(CAMPOS_FIJOS)}
    for j in range(6, len(partes)):
        resultado[f"BLOQUE_{j + 1}"] = partes[j]
    return resultado


def iter_registros(base: Path) -> Generator[dict, None, None]:
    """
    Generador: produce un dict por archivo de audio encontrado.
    NO carga todo en memoria a la vez.
    """
    mes = base.name
    for carpeta_dia in sorted(base.iterdir()):
        if not carpeta_dia.is_dir():
            continue
        dia = carpeta_dia.name
        for carpeta_indice in sorted(carpeta_dia.iterdir()):
            if not carpeta_indice.is_dir():
                continue
            indice = carpeta_indice.name
            for archivo in sorted(carpeta_indice.iterdir()):
                if not archivo.is_file():
                    continue
                if archivo.suffix.lower() not in EXTENSIONES:
                    continue
                campos = parsear_nombre(archivo.stem)
                registro = {
                    "MES"            : mes,
                    "DIA"            : dia,
                    "INDICE"         : indice,
                    "YEAR"           : campos.get("YEAR", ""),
                    "COD1"           : campos.get("COD1", ""),
                    "AoP"            : campos.get("AoP", ""),
                    "COD2"           : campos.get("COD2", ""),
                    "COD3"           : campos.get("COD3", ""),
                    "TELEFONO"       : campos.get("TELEFONO", ""),
                    "PESO"           : formatear_peso(archivo.stat().st_size),
                    "RUTA"           : str(archivo.resolve()),
                    "NOMBRE_COMPLETO": archivo.name,
                }
                for clave, valor in campos.items():
                    if clave.startswith("BLOQUE_"):
                        registro[clave] = valor
                yield registro


# ── Pre-scan para detectar bloques extra ───────────────────────────────────────

def detectar_bloques_extra(base: Path) -> list[str]:
    """
    Hace un primer paso rápido (solo nombres de archivo, sin stat())
    para descubrir qué BLOQUE_N existen en el lote.
    """
    extras: set[str] = set()
    for carpeta_dia in base.iterdir():
        if not carpeta_dia.is_dir():
            continue
        for carpeta_indice in carpeta_dia.iterdir():
            if not carpeta_indice.is_dir():
                continue
            for archivo in carpeta_indice.iterdir():
                if not archivo.is_file():
                    continue
                if archivo.suffix.lower() not in EXTENSIONES:
                    continue
                partes = archivo.stem.split("-")
                for j in range(6, len(partes)):
                    extras.add(f"BLOQUE_{j + 1}")
    return sorted(extras, key=lambda x: int(x.split("_")[1]))


# ── Exportar ──────────────────────────────────────────────────────────────────

def exportar_en_partes(base: Path, salida_base: Path):
    """
    Escribe uno o varios archivos Excel dividiendo los registros en chunks
    de MAX_FILAS_POR_ARCHIVO para respetar el límite de filas de Excel.
    Usa Workbook estándar (no write_only) para mantener column_dimensions
    y freeze_panes con estilos completos.
    Devuelve (lista_de_rutas, total_registros).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n[ERROR] La librería 'openpyxl' no está instalada.")
        print("  Ejecuta:  pip install openpyxl\n")
        sys.exit(1)

    # ── Estilos (creados una sola vez, reutilizados en todos los archivos) ─────
    COLOR_HEADER   = "1A3A5C"
    COLOR_EXTRA    = "8B4513"
    COLOR_FILA_PAR = "EAF1FB"

    f_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    f_dato   = Font(name="Calibri", size=10)
    f_total  = Font(name="Calibri", bold=True, size=10)

    fill_hdr   = PatternFill("solid", fgColor=COLOR_HEADER)
    fill_extra = PatternFill("solid", fgColor=COLOR_EXTRA)
    fill_par   = PatternFill("solid", fgColor=COLOR_FILA_PAR)
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    borde = Border(
        left   = Side(style="thin", color="BFBFBF"),
        right  = Side(style="thin", color="BFBFBF"),
        top    = Side(style="thin", color="BFBFBF"),
        bottom = Side(style="thin", color="BFBFBF"),
    )
    a_centro = Alignment(horizontal="center", vertical="center")
    a_izq    = Alignment(horizontal="left",   vertical="center")

    # ── Detectar columnas extra ────────────────────────────────────────────────
    print("  [1/3] Analizando nombres de archivos para detectar columnas extra...")
    bloques_extra  = detectar_bloques_extra(base)
    columnas       = COLUMNAS_FIJAS + [(b, 14) for b in bloques_extra]
    nombres_col    = [c[0] for c in columnas]
    anchos_col     = [c[1] for c in columnas]
    columnas_centradas = COLUMNAS_CENTRADAS_FIJAS | set(bloques_extra)

    if bloques_extra:
        print(f"     → Bloques extra detectados: {', '.join(bloques_extra)}")

    # ── Helpers ────────────────────────────────────────────────────────────────
    def inicializar_hoja(ws) -> None:
        """Aplica anchos, freeze_panes y fila de cabecera a una hoja activa."""
        for idx, (nom, ancho) in enumerate(zip(nombres_col, anchos_col), start=1):
            ws.column_dimensions[get_column_letter(idx)].width = ancho
            celda = ws.cell(row=1, column=idx, value=nom)
            celda.font      = f_header
            celda.fill      = fill_extra if nom not in NOMBRES_FIJOS else fill_hdr
            celda.alignment = a_centro
            celda.border    = borde
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    def escribir_fila(ws, fila_excel: int, registro: dict) -> None:
        """Escribe una fila de datos con estilos."""
        fila_par = (fila_excel % 2 == 0)
        fill = fill_par if fila_par else fill_impar
        for col_idx, nom in enumerate(nombres_col, start=1):
            celda = ws.cell(row=fila_excel, column=col_idx,
                            value=registro.get(nom, ""))
            celda.font      = f_dato
            celda.fill      = fill
            celda.border    = borde
            celda.alignment = a_centro if nom in columnas_centradas else a_izq

    def guardar_parte(wb, ruta: Path, n_filas: int) -> None:
        ruta.parent.mkdir(parents=True, exist_ok=True)
        wb.save(ruta)
        print(f"     → Guardado: {ruta}  ({n_filas:,} filas)")

    # ── Iterar registros y escribir en chunks ──────────────────────────────────
    print("  [2/3] Escribiendo registros...")

    stem   = salida_base.stem
    sufijo = salida_base.suffix

    archivos_generados: list[Path] = []
    parte           = 1
    fila_excel      = 2          # fila 1 = cabecera
    filas_en_parte  = 0
    total_registros = 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    inicializar_hoja(ws)

    for registro in iter_registros(base):
        # ¿Se llenó la parte actual?
        if filas_en_parte >= MAX_FILAS_POR_ARCHIVO:
            ruta_parte = salida_base.parent / f"{stem}_parte_{parte:03d}{sufijo}"
            guardar_parte(wb, ruta_parte, filas_en_parte)
            archivos_generados.append(ruta_parte)
            parte          += 1
            fila_excel      = 2
            filas_en_parte  = 0
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            inicializar_hoja(ws)

        escribir_fila(ws, fila_excel, registro)
        fila_excel      += 1
        filas_en_parte  += 1
        total_registros += 1

        if total_registros % 50_000 == 0:
            print(f"     → {total_registros:,} registros procesados...")

    # ── Guardar última (o única) parte ────────────────────────────────────────
    if filas_en_parte > 0:
        # Fila resumen al final
        celda_tot = ws.cell(row=fila_excel, column=1,
                            value=f"Total registros en esta parte: {filas_en_parte:,}")
        celda_tot.font      = f_total
        celda_tot.alignment = a_izq

        ruta_final = salida_base if parte == 1 else \
                     salida_base.parent / f"{stem}_parte_{parte:03d}{sufijo}"
        guardar_parte(wb, ruta_final, filas_en_parte)
        archivos_generados.append(ruta_final)

    return archivos_generados, total_registros


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Generador de Reporte de Audios")
    print("=" * 60)
    print(f"  Carpeta base   : {BASE_FOLDER}")
    print(f"  Archivo salida : {SALIDA_EXCEL}")
    print(f"  Máx filas/xlsx : {MAX_FILAS_POR_ARCHIVO:,}")
    print()

    if not BASE_FOLDER.exists():
        print(f"[ERROR] La carpeta base no existe:\n  {BASE_FOLDER}")
        sys.exit(1)
    if not BASE_FOLDER.is_dir():
        print(f"[ERROR] La ruta indicada no es una carpeta:\n  {BASE_FOLDER}")
        sys.exit(1)

    archivos, total = exportar_en_partes(BASE_FOLDER, SALIDA_EXCEL)

    print()
    print("=" * 60)
    print(f"  [3/3] Completado")
    print(f"  Total registros : {total:,}")
    print(f"  Archivos Excel  : {len(archivos)}")
    for a in archivos:
        print(f"    · {a}")
    print()
    print("  Columnas fijas : MES | DIA | INDICE | YEAR | COD1 | AoP |")
    print("                   COD2 | COD3 | TELEFONO | PESO | RUTA | NOMBRE_COMPLETO")
    if len(archivos) > 1:
        print(f"\n  NOTA: Los datos se dividieron en {len(archivos)} archivos")
        print("        porque Excel tiene un límite de 1 048 576 filas.")
    print("=" * 60)


if __name__ == "__main__":
    main()
