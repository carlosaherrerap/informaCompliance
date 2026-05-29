import pyodbc
import json
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

# ==========================================
# CONFIGURACIÓN DEL SCRIPT
# ==========================================
# Nombre del Excel de donde se leerán los nombres (Asegúrate de crearlo o cambiar el nombre aquí)
INPUT_EXCEL = "nombres_input.xlsx" 
# Nombre del Excel donde se guardarán los resultados extraídos
OUTPUT_EXCEL = "resultados_reniec.xlsx"
# Archivo de control para reanudar el script si se corta
STATE_FILE = "estado_avance.json"

# Tamaño de lote (registros a consultar al mismo tiempo)
# 100 es ideal para equilibrar velocidad y límites de parámetros en SQL Server (max 2100 params)
CHUNK_SIZE = 100 

DB_CONN_STR = (
    "DRIVER={SQL Server};"
    "SERVER=192.168.1.42;"
    "DATABASE=DataStage;"
    "UID=sa;"
    "PWD=Administrador2025$$;"
)

def get_db_connection():
    return pyodbc.connect(DB_CONN_STR)

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r') as f:
            return json.load(f).get('last_index', 0)
    return 0

def save_state(index):
    with open(STATE_FILE, 'w') as f:
        json.dump({'last_index': index}, f)

def main():
    print(f"--- SCRIPT DE EXTRACCIÓN MASIVA RENIEC ---")
    if not os.path.exists(INPUT_EXCEL):
        print(f"Error: No se encontró el archivo de entrada '{INPUT_EXCEL}'.")
        print("Por favor, asegúrate de tener este archivo con las columnas obligatorias: NOMBRES, APE_PAT, APE_MAT")
        return

    print("Cargando archivo Excel de entrada en memoria...")
    try:
        df_input = pd.read_excel(INPUT_EXCEL)
    except Exception as e:
        print(f"Error al leer el archivo Excel de entrada: {e}")
        return

    required_cols = ['NOMBRES', 'APE_PAT', 'APE_MAT']
    for col in required_cols:
        if col not in df_input.columns:
            print(f"Error Crítico: La columna '{col}' no existe en {INPUT_EXCEL}")
            return

    # Limpiar datos: rellenar celdas vacías, convertir a texto y eliminar espacios en blanco
    for col in required_cols:
        df_input[col] = df_input[col].fillna('').astype(str).str.strip()

    start_idx = load_state()
    total_rows = len(df_input)
    
    if start_idx >= total_rows:
        print("El procesamiento ya ha sido completado al 100%. Para reiniciar, borra el archivo 'estado_avance.json'.")
        return

    print(f"Conectando a la base de datos SQL Server (192.168.1.42)...")
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
    except Exception as e:
        print(f"Error de conexión a la BD: {e}")
        return

    print(f"Iniciando desde la fila {start_idx} de un total de {total_rows}...")

    # Abrir el Excel existente o crear uno nuevo si apenas empieza
    if os.path.exists(OUTPUT_EXCEL):
        print(f"Abriendo archivo de resultados existente '{OUTPUT_EXCEL}' para continuar agregando datos...")
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active
        columns_initialized = True
    else:
        print(f"Creando nuevo archivo de resultados '{OUTPUT_EXCEL}'...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        columns_initialized = False

    # Extraer nombres de las columnas para ponerlas en la cabecera del Excel (Solo la primera vez)
    if not columns_initialized:
        try:
            # Consulta ultra-rápida (TOP 0) solo para traer los metadatos de las columnas de la tabla
            cursor.execute("SELECT TOP 0 * FROM Reniec.Base")
            columns = [column[0] for column in cursor.description]
            ws.append(columns)
            wb.save(OUTPUT_EXCEL)
            columns_initialized = True
        except Exception as e:
            print(f"No se pudo inicializar las columnas leyendo Reniec.Base: {e}")
            return

    # Bucle principal: Procesamos fila por fila para asegurar el uso de índices de SQL Server
    # y guardamos el Excel cada CHUNK_SIZE (100) registros.
    coincidencias_lote = 0
    
    for i in range(start_idx, total_rows):
        row = df_input.iloc[i]
        nombres = row['NOMBRES']
        ape_pat = row['APE_PAT']
        ape_mat = row['APE_MAT']
        
        query = "SELECT * FROM Reniec.Base WHERE PRENOMBRES=? AND APE_PAT=? AND APE_MAT=?"
        
        try:
            cursor.execute(query, (nombres, ape_pat, ape_mat))
            results = cursor.fetchall()
            
            # Si hay resultados, los agregamos directamente al Excel
            if results:
                for res in results:
                    cleaned_row = [val if val is not None else "" for val in res]
                    ws.append(cleaned_row)
                coincidencias_lote += len(results)
                
            # Log por cada registro para saber exactamente en dónde vamos
            print(f"[{i + 1}/{total_rows}] Consultando: {nombres} {ape_pat} {ape_mat} -> Encontrados: {len(results)}")
            
            # Guardar progreso físico cada 100 registros o si es el último registro
            if (i + 1) % CHUNK_SIZE == 0 or (i + 1) == total_rows:
                wb.save(OUTPUT_EXCEL)
                save_state(i + 1)
                print(f"\n--- [GUARDADO AUTOMÁTICO] Progreso guardado hasta la fila {i + 1}. Coincidencias en este lote: {coincidencias_lote} ---\n")
                coincidencias_lote = 0
                
        except Exception as e:
            print(f"\n[ERROR] Problema procesando la fila {i + 1} ({nombres} {ape_pat}): {e}")
            print("El estado de avance está guardado. Cuando soluciones el error, vuelve a correr el script y continuará desde aquí.")
            break

    conn.close()
    print("\nProceso finalizado exitosamente. Revisa el archivo 'resultados_reniec.xlsx'.")

if __name__ == "__main__":
    main()
